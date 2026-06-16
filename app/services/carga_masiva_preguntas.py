"""Carga masiva de preguntas desde un archivo Excel (.xlsx).

Formato esperado (una fila por pregunta, primera fila = encabezados):

    evento | grupo | seccion | tipo_pregunta | pregunta |
    opcion_1 | opcion_2 | opcion_3 | opcion_4 | opcion_correcta

- evento: nombre del evento (debe existir).
- grupo: nombre del grupo (opcional). Vacío => sección común del evento;
  con valor => sección específica de ese grupo (el grupo debe existir en el evento).
- seccion: nombre de la sección (se crea si no existe, con el alcance según 'grupo').
- tipo_pregunta: 'opcion_unica' | 'abierta' | 'opcion_opinion'.
- opcion_1..4: textos de las opciones (vacío para 'abierta').
- opcion_correcta: número 1-4 (solo 'opcion_unica'; vacío en los demás).

El procesamiento es por fila; los errores no detienen la carga: se acumulan y se
reportan. Cada pregunta válida se inserta con sus respuestas en su propia transacción.
"""
from io import BytesIO
from typing import Any, Optional

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.evento import Evento
from app.models.grupo import Grupo
from app.models.seccion import Seccion
from app.models.pregunta import Pregunta
from app.models.respuesta import Respuesta

ENCABEZADOS = [
    "evento", "grupo", "seccion", "tipo_pregunta", "pregunta",
    "opcion_1", "opcion_2", "opcion_3", "opcion_4", "opcion_correcta",
]

TIPOS_VALIDOS = {"opcion_unica", "abierta", "opcion_opinion"}
TIPOS_CON_OPCIONES = {"opcion_unica", "opcion_opinion"}


class FilaError(Exception):
    """Error de validación de una fila concreta."""


def _norm(value: Any) -> str:
    """Normaliza una celda a string limpio ('' si es None)."""
    if value is None:
        return ""
    return str(value).strip()


def parse_excel(contenido: bytes) -> list[dict]:
    """Lee el .xlsx y devuelve una lista de dicts (una por fila de datos).

    Valida que los encabezados esperados estén presentes.
    """
    wb = load_workbook(filename=BytesIO(contenido), read_only=True, data_only=True)
    ws = wb.active
    filas = ws.iter_rows(values_only=True)

    try:
        cabecera = [_norm(c).lower() for c in next(filas)]
    except StopIteration as exc:
        raise ValueError("El archivo está vacío") from exc

    faltantes = [h for h in ENCABEZADOS if h not in cabecera]
    if faltantes:
        raise ValueError(f"Faltan columnas en el Excel: {', '.join(faltantes)}")

    idx = {h: cabecera.index(h) for h in ENCABEZADOS}
    resultado = []
    for n, fila in enumerate(filas, start=2):  # fila 1 = encabezados
        if fila is None or all(_norm(c) == "" for c in fila):
            continue  # saltar filas en blanco
        resultado.append({
            "_fila": n,
            **{h: _norm(fila[idx[h]]) if idx[h] < len(fila) else "" for h in ENCABEZADOS},
        })
    return resultado


def _opciones_de_fila(fila: dict) -> list[dict]:
    """Extrae las opciones no vacías como [{respuesta, orden}], en orden 1..n."""
    opciones = []
    orden = 1
    for col in ("opcion_1", "opcion_2", "opcion_3", "opcion_4"):
        texto = fila.get(col, "")
        if texto:
            opciones.append({"respuesta": texto, "orden": orden})
            orden += 1
    return opciones


async def _resolver_seccion(
    db: AsyncSession, id_evento: int, id_grupo: Optional[int], nombre_seccion: str
) -> Seccion:
    """Busca la sección por (evento, grupo, nombre); la crea si no existe."""
    cond = [Seccion.id_evento == id_evento, Seccion.nombre_seccion == nombre_seccion]
    cond.append(Seccion.id_grupo.is_(None) if id_grupo is None else Seccion.id_grupo == id_grupo)
    seccion = (await db.execute(select(Seccion).where(*cond))).scalar_one_or_none()
    if seccion is None:
        seccion = Seccion(
            id_evento=id_evento, id_grupo=id_grupo, nombre_seccion=nombre_seccion
        )
        db.add(seccion)
        await db.flush()
    return seccion


async def _validar_fila(db: AsyncSession, fila: dict) -> dict:
    """Valida una fila y devuelve los datos resueltos para insertar.

    Lanza FilaError con un mensaje legible si algo no cuadra.
    """
    tipo = fila["tipo_pregunta"]
    if tipo not in TIPOS_VALIDOS:
        raise FilaError(f"tipo_pregunta inválido: '{tipo}'")
    if not fila["pregunta"]:
        raise FilaError("La pregunta está vacía")
    if not fila["evento"]:
        raise FilaError("Falta el evento")
    if not fila["seccion"]:
        raise FilaError("Falta la sección")

    # Evento
    evento = (await db.execute(
        select(Evento).where(Evento.nombre_evento == fila["evento"])
    )).scalar_one_or_none()
    if evento is None:
        raise FilaError(f"El evento '{fila['evento']}' no existe")

    # Grupo (opcional)
    id_grupo = None
    if fila["grupo"]:
        grupo = (await db.execute(
            select(Grupo).where(
                Grupo.id_evento == evento.id_evento,
                Grupo.nombre_grupo == fila["grupo"],
            )
        )).scalar_one_or_none()
        if grupo is None:
            raise FilaError(
                f"El grupo '{fila['grupo']}' no existe en el evento '{fila['evento']}'"
            )
        id_grupo = grupo.id_grupo

    # Opciones y opción correcta según tipo
    opciones = _opciones_de_fila(fila)
    opcion_correcta = None
    if tipo in TIPOS_CON_OPCIONES:
        if not (1 <= len(opciones) <= 4):
            raise FilaError("Debe haber entre 1 y 4 opciones")
        if tipo == "opcion_unica":
            if not fila["opcion_correcta"]:
                raise FilaError("Falta opcion_correcta para opcion_unica")
            try:
                opcion_correcta = int(float(fila["opcion_correcta"]))
            except ValueError as exc:
                raise FilaError("opcion_correcta debe ser un número") from exc
            if opcion_correcta not in [o["orden"] for o in opciones]:
                raise FilaError("opcion_correcta no corresponde a ninguna opción")
        elif fila["opcion_correcta"]:
            raise FilaError("opcion_opinion no debe tener opcion_correcta")
    else:  # abierta
        if opciones:
            raise FilaError("Las preguntas abiertas no deben tener opciones")
        if fila["opcion_correcta"]:
            raise FilaError("Las preguntas abiertas no deben tener opcion_correcta")

    return {
        "id_evento": evento.id_evento,
        "id_grupo": id_grupo,
        "nombre_seccion": fila["seccion"],
        "pregunta": fila["pregunta"],
        "tipo_pregunta": tipo,
        "opciones": opciones,
        "opcion_correcta": opcion_correcta,
    }


async def cargar_preguntas(db: AsyncSession, contenido: bytes) -> dict:
    """Procesa el Excel y crea las preguntas. Devuelve un resumen con errores por fila."""
    filas = parse_excel(contenido)
    creadas = 0
    errores: list[dict] = []

    for fila in filas:
        n = fila["_fila"]
        try:
            datos = await _validar_fila(db, fila)
            seccion = await _resolver_seccion(
                db, datos["id_evento"], datos["id_grupo"], datos["nombre_seccion"]
            )
            nueva = Pregunta(
                id_seccion=seccion.id_seccion,
                pregunta=datos["pregunta"],
                tipo_pregunta=datos["tipo_pregunta"],
                opcion_correcta=datos["opcion_correcta"],
            )
            db.add(nueva)
            await db.flush()
            for op in datos["opciones"]:
                db.add(Respuesta(
                    id_pregunta=nueva.id_pregunta,
                    orden=op["orden"],
                    respuesta=op["respuesta"],
                ))
            await db.commit()
            creadas += 1
        except FilaError as exc:
            await db.rollback()
            errores.append({"fila": n, "motivo": str(exc)})
        except IntegrityError:
            await db.rollback()
            errores.append({
                "fila": n,
                "motivo": "Pregunta duplicada en la sección (ya existe ese texto)",
            })

    return {
        "total_filas": len(filas),
        "creadas": creadas,
        "con_error": len(errores),
        "errores": errores,
    }
