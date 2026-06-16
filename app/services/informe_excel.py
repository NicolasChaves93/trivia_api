"""Generación del informe de resultados en Excel (.xlsx) de 4 hojas.

Hojas:
  1. 'Respuestas G'   -> métricas generales + ranking por participación.
  2. 'Respuestas I'   -> detalle por usuario x pregunta (incluye opinión/abiertas aparte).
  3. 'Grafica Respuesta' -> pivote por sección/pregunta (correctas/incorrectas/total).
  4. 'Graficas Puntuacion' -> histograma de nº de correctas por usuario.

Manejo de intentos (`numero_intento`) configurable:
  - 'todos'   -> una fila por intento finalizado.
  - 'primero' -> solo el primer intento de cada usuario en el grupo.
  - 'ultimo'  -> solo el último intento.
  - 'mejor'   -> el intento con más correctas (desempata por menor tiempo).

Las preguntas de opinión ('opcion_opinion') y abiertas NO puntúan: se excluyen de
correctas/incorrectas/%, pero se listan en la hoja de detalle.
"""
from io import BytesIO
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.participacion import Participacion, EstadoParticipacion
from app.models.resultado import Resultado
from app.models.usuario import Usuario
from app.models.grupo import Grupo
from app.models.evento import Evento
from app.models.seccion import Seccion
from app.models.pregunta import Pregunta
from app.models.respuesta import Respuesta
from app.models.respuesta_usuario import RespuestaUsuario

INTENTOS_VALIDOS = {"todos", "primero", "ultimo", "mejor"}
_HDR_FILL = PatternFill("solid", fgColor="305496")
_HDR_FONT = Font(bold=True, color="FFFFFF")


def _fmt_tiempo(td) -> str:
    """Formatea un timedelta a HH:MM:SS."""
    if td is None:
        return ""
    total = int(td.total_seconds())
    return f"{total // 3600:02d}:{(total % 3600) // 60:02d}:{total % 60:02d}"


def _estilo_encabezado(ws, fila: int = 1) -> None:
    for c in ws[fila]:
        if c.value is not None:
            c.font = _HDR_FONT
            c.fill = _HDR_FILL


def _filtrar_intentos(participaciones: list, modo: str) -> list:
    """Aplica el filtro de intentos sobre (Participacion, Resultado, Usuario, Grupo)."""
    if modo == "todos":
        return participaciones
    por_usuario: dict = {}
    for fila in participaciones:
        part = fila[0]
        clave = (part.id_usuario, part.id_grupo)
        por_usuario.setdefault(clave, []).append(fila)
    elegidas = []
    for filas in por_usuario.values():
        if modo == "primero":
            elegidas.append(min(filas, key=lambda f: f[0].numero_intento))
        elif modo == "ultimo":
            elegidas.append(max(filas, key=lambda f: f[0].numero_intento))
        else:  # mejor
            elegidas.append(max(
                filas,
                key=lambda f: (f[1].respuestas_correctas, -(f[1].tiempo_total.total_seconds()
                                                             if f[1].tiempo_total else 0)),
            ))
    return elegidas


async def generar_informe_excel(
    db: AsyncSession,
    id_evento: int,
    id_grupo: Optional[int] = None,
    intentos: str = "todos",
) -> bytes:
    """Genera el informe .xlsx y lo devuelve como bytes."""
    if intentos not in INTENTOS_VALIDOS:
        raise ValueError(f"intentos debe ser uno de {sorted(INTENTOS_VALIDOS)}")

    # --- Participaciones finalizadas (con resultado) ---
    stmt = (
        select(Participacion, Resultado, Usuario, Grupo, Evento)
        .join(Resultado, Resultado.id_participacion == Participacion.id_participacion)
        .join(Usuario, Usuario.id_usuario == Participacion.id_usuario)
        .join(Grupo, Grupo.id_grupo == Participacion.id_grupo)
        .join(Evento, Evento.id_evento == Grupo.id_evento)
        .where(Grupo.id_evento == id_evento,
               Participacion.estado == EstadoParticipacion.FINALIZADO)
    )
    if id_grupo is not None:
        stmt = stmt.where(Participacion.id_grupo == id_grupo)
    filas = (await db.execute(stmt)).all()
    filas = _filtrar_intentos(filas, intentos)

    # Ranking: más correctas, menos tiempo
    filas.sort(key=lambda f: (-f[1].respuestas_correctas,
                              f[1].tiempo_total.total_seconds() if f[1].tiempo_total else 0))
    ids_part = [f[0].id_participacion for f in filas]

    # --- Detalle de respuestas (opción única + opinión) ---
    detalle = []
    if ids_part:
        det_stmt = (
            select(
                Grupo.nombre_grupo, Seccion.nombre_seccion, Usuario.cedula, Usuario.nombre,
                Participacion.numero_intento, Pregunta.id_pregunta, Pregunta.pregunta,
                Pregunta.tipo_pregunta, Pregunta.opcion_correcta, RespuestaUsuario.orden_seleccionado,
            )
            .join(Participacion, Participacion.id_participacion == RespuestaUsuario.id_participacion)
            .join(Pregunta, Pregunta.id_pregunta == RespuestaUsuario.id_pregunta)
            .join(Seccion, Seccion.id_seccion == Pregunta.id_seccion)
            .join(Usuario, Usuario.id_usuario == Participacion.id_usuario)
            .join(Grupo, Grupo.id_grupo == Participacion.id_grupo)
            .where(RespuestaUsuario.id_participacion.in_(ids_part))
        )
        detalle = (await db.execute(det_stmt)).all()

        # Mapa (id_pregunta, orden) -> texto de respuesta
        ids_preg = {d[5] for d in detalle}
        textos = {}
        if ids_preg:
            for r in (await db.execute(
                select(Respuesta).where(Respuesta.id_pregunta.in_(ids_preg))
            )).scalars():
                textos[(r.id_pregunta, r.orden)] = r.respuesta

    wb = Workbook()

    # ===== Hoja 1: Respuestas G =====
    g = wb.active
    g.title = "Respuestas G"
    total_correctas = sum(f[1].respuestas_correctas for f in filas)
    total_incorrectas = sum(f[1].respuestas_incorrectas for f in filas)
    total_resp = total_correctas + total_incorrectas
    g.append(["MÉTRICAS GENERALES"])
    g.append(["Total Participaciones", len(filas)])
    g.append(["Total Respuestas (puntúan)", total_resp])
    g.append(["Correctas", total_correctas,
              round(total_correctas / total_resp, 4) if total_resp else 0])
    g.append(["Incorrectas", total_incorrectas,
              round(total_incorrectas / total_resp, 4) if total_resp else 0])
    g.append([])
    encabezado_g = ["Evento", "Ranking", "Cédula", "Nombre", "Grupo", "Intento",
                    "Total", "Correctas", "Incorrectas", "Acierto %", "Tiempo", "Fecha Inicio"]
    g.append(encabezado_g)
    fila_hdr_g = g.max_row
    for i, (part, res, usr, grp, ev) in enumerate(filas, start=1):
        g.append([
            ev.nombre_evento, i, usr.cedula, usr.nombre, grp.nombre_grupo, part.numero_intento,
            res.total_preguntas, res.respuestas_correctas, res.respuestas_incorrectas,
            float(res.porcentaje_acierto), _fmt_tiempo(res.tiempo_total),
            part.started_at.replace(tzinfo=None) if part.started_at else None,
        ])
    _estilo_encabezado(g, fila_hdr_g)

    # ===== Hoja 2: Respuestas I (detalle) =====
    di = wb.create_sheet("Respuestas I")
    di.append(["Grupo", "Sección", "Cédula", "Nombre", "Intento", "Pregunta",
               "Enunciado", "Tipo", "Selección", "Respuesta Usuario",
               "Opción Correcta", "Respuesta Correcta", "Resultado"])
    for (grupo_n, secc, ced, nom, intento, idp, enun, tipo, correcta, sel) in detalle:
        texto_sel = textos.get((idp, sel), "")
        if tipo == "opcion_unica":
            texto_corr = textos.get((idp, correcta), "")
            resultado = "Correcta" if sel == correcta else "Incorrecta"
            corr_val = correcta
        else:  # opcion_opinion (no puntúa)
            texto_corr = ""
            resultado = "Opinión"
            corr_val = ""
        di.append([grupo_n, secc, ced, nom, intento, idp, enun, tipo,
                   sel, texto_sel, corr_val, texto_corr, resultado])
    _estilo_encabezado(di)

    # ===== Hoja 3: Grafica Respuesta (pivote por sección/pregunta) =====
    gr = wb.create_sheet("Grafica Respuesta")
    gr.append(["Sección", "Pregunta", "Enunciado", "Correcta", "Incorrecta", "Total"])
    agg: dict = {}
    orden_preg = []
    for (_g, secc, _c, _n, _i, idp, enun, tipo, correcta, sel) in detalle:
        if tipo != "opcion_unica":
            continue  # solo puntúan opción única
        k = (secc, idp, enun)
        if k not in agg:
            agg[k] = [0, 0]
            orden_preg.append(k)
        if sel == correcta:
            agg[k][0] += 1
        else:
            agg[k][1] += 1
    for (secc, idp, enun) in orden_preg:
        c, inc = agg[(secc, idp, enun)]
        gr.append([secc, idp, enun, c, inc, c + inc])
    _estilo_encabezado(gr)

    # ===== Hoja 4: Graficas Puntuacion (histograma de correctas) =====
    gp = wb.create_sheet("Graficas Puntuacion")
    gp.append(["Total respuestas correctas por usuario", "No. Usuarios", "% Usuarios"])
    hist: dict = {}
    for f in filas:
        hist[f[1].respuestas_correctas] = hist.get(f[1].respuestas_correctas, 0) + 1
    total_u = len(filas)
    for correctas in sorted(hist.keys(), reverse=True):
        n = hist[correctas]
        gp.append([correctas, n, round(n / total_u, 4) if total_u else 0])
    _estilo_encabezado(gp)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
